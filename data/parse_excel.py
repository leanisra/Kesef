import openpyxl
from datetime import datetime, date
import uuid
import os

ADMIN  = r"C:\Users\Leandro\Downloads\241227_Guatemala 5934_ admin (2).xlsx"
CUOTAS = r"C:\Users\Leandro\Downloads\Guatemala 5934_ cuotas inversores (2).xlsx"
OUTPUT = r"C:\Users\Leandro\Desktop\kesef\data"

os.makedirs(OUTPUT, exist_ok=True)

# ── helpers ──────────────────────────────────────────────────────────────────

def q(v):
    if v is None:
        return 'NULL'
    if isinstance(v, bool):
        return 'TRUE' if v else 'FALSE'
    if isinstance(v, (int, float)):
        r = round(v, 4)
        return str(int(r)) if r == int(r) else str(r)
    return "'" + str(v).replace("'", "''") + "'"

def fmt_date(v):
    if v is None:
        return 'NULL'
    if isinstance(v, datetime):
        return "'" + v.strftime('%Y-%m-%d') + "'"
    if isinstance(v, date):
        return "'" + v.isoformat() + "'"
    if isinstance(v, str):
        s = v.strip()
        for fmt in ['%d-%m-%y', '%d/%m/%y', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d']:
            try:
                return "'" + datetime.strptime(s, fmt).strftime('%Y-%m-%d') + "'"
            except Exception:
                pass
    return 'NULL'

def is_num(v):
    return isinstance(v, (int, float))

def cobrado_in_row(row):
    # 'pagado' sola puede ser encabezado de columna (junto a 'saldo'); usamos 'cobrado' y 'pagada'
    for c in row:
        if isinstance(c, str) and c.strip().lower() in ('cobrado', 'pagada'):
            return True
    return False

def gen_uuid():
    return str(uuid.uuid4())

# ── pre-generate UUIDs ────────────────────────────────────────────────────────
C_801 = gen_uuid(); C_402 = gen_uuid(); C_EGA = gen_uuid()
C_202 = gen_uuid(); C_401 = gen_uuid()

T_801 = gen_uuid()
T_402 = gen_uuid()
T_201 = gen_uuid(); T_301 = gen_uuid(); T_501 = gen_uuid(); T_701 = gen_uuid()
T_202 = gen_uuid()
T_401 = gen_uuid()

CAJA_CASH  = gen_uuid()
CAJA_BANCO = gen_uuid()
CAJA_MN    = gen_uuid()

# ── cuota base data (extracted from Excel headers) ────────────────────────────
CUOTA_A_801 = 2781430.208; CUOTA_B_801 = 6490003.819
TOTAL_USD_801 = 509575.0; ANTICIPO_USD_801 = 254787.5
TC_801 = 1310; SPLIT_A_801 = 50; N_CUOTAS_801 = 12

CUOTA_A_402 = 2142244.65; CUOTA_B_402 = 2999142.51
TOTAL_USD_402 = 247777.69; ANTICIPO_USD_402 = 99111.08
TC_402 = 1245; SPLIT_A_402 = 40; N_CUOTAS_402 = 12

CUOTA_A_202 = 7487688.333; CUOTA_B_202 = 1497537.667
TOTAL_USD_202 = 289846.0; ANTICIPO_USD_202 = 115938.4
TC_202 = 1550; SPLIT_A_202 = 50; N_CUOTAS_202 = 6

CUOTA_A_401 = 1174666.32; CUOTA_B_401 = 5873331.6
TOTAL_USD_401 = 261036.96; ANTICIPO_USD_401 = 104414.784
TC_401 = 1080; SPLIT_A_401 = 50; N_CUOTAS_401 = 2

EGA_DATA = {
    '201': {'cuota_a': 6824382.5,   'cuota_b': 1364876.5,   'total_usd': 277602.0,    'anticipo': 111040.8,  'n': 6},
    '301': {'cuota_a': 6789479.305, 'cuota_b': 1357895.861, 'total_usd': 276182.209,  'anticipo': 110472.88, 'n': 6},
    '501': {'cuota_a': 7715282.5,   'cuota_b': 1543056.5,   'total_usd': 313842.0,    'anticipo': 125536.8,  'n': 6},
    '701': {'cuota_a': 8359636.68,  'cuota_b': 1671927.336, 'total_usd': 340053.0175, 'anticipo': 136021.207,'n': 6},
}
TC_EGA = 1475; SPLIT_A_EGA = 50

# ── 1. CLIENTES SQL ───────────────────────────────────────────────────────────
clientes_lines = [
    "-- insertar_clientes.sql",
    "-- Clientes de la obra Guatemala 5934",
    "",
    "INSERT INTO clientes (id, nombre, telefono, email) VALUES",
    "  (" + q(C_801) + ", 'Sharon Menajovsky y Matias Bekerman', NULL, NULL),",
    "  (" + q(C_402) + ", 'Ruth Fischer y Beny', NULL, NULL),",
    "  (" + q(C_EGA) + ", 'Edgardo Aguirre', NULL, NULL),",
    "  (" + q(C_202) + ", 'Gelles', NULL, NULL),",
    "  (" + q(C_401) + ", 'Alberto Israel y Gabriela Hirschl', NULL, NULL)",
    "ON CONFLICT (id) DO NOTHING;",
]
with open(os.path.join(OUTPUT, 'insertar_clientes.sql'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(clientes_lines))
print("insertar_clientes.sql generado")

# ── 2. CONTRATOS SQL ──────────────────────────────────────────────────────────
# Contrato state: vigente for all
OBRA_QUERY = "(SELECT id FROM obras WHERE nombre LIKE '%Guatemala%' LIMIT 1)"

contratos_lines = [
    "-- insertar_contratos.sql",
    "-- Contratos de la obra Guatemala 5934",
    "-- IMPORTANTE: correr insertar_clientes.sql antes",
    "",
    "DO $$ DECLARE obra_id UUID;",
    "BEGIN",
    "  SELECT id INTO obra_id FROM obras WHERE nombre LIKE '%Guatemala%' LIMIT 1;",
    "",
    "  -- Insertar contratos",
    "  INSERT INTO contratos (id, obra_id, cliente_id, estado, precio_usd_total, anticipo_usd, split_a_pct, n_cuotas, tc_contrato) VALUES",
]

contrato_rows = [
    (T_801, C_801, 'vigente', TOTAL_USD_801, ANTICIPO_USD_801, SPLIT_A_801, N_CUOTAS_801, TC_801),
    (T_402, C_402, 'vigente', TOTAL_USD_402, ANTICIPO_USD_402, SPLIT_A_402, N_CUOTAS_402, TC_402),
    (T_201, C_EGA, 'vigente', EGA_DATA['201']['total_usd'], EGA_DATA['201']['anticipo'], SPLIT_A_EGA, EGA_DATA['201']['n'], TC_EGA),
    (T_301, C_EGA, 'vigente', EGA_DATA['301']['total_usd'], EGA_DATA['301']['anticipo'], SPLIT_A_EGA, EGA_DATA['301']['n'], TC_EGA),
    (T_501, C_EGA, 'vigente', EGA_DATA['501']['total_usd'], EGA_DATA['501']['anticipo'], SPLIT_A_EGA, EGA_DATA['501']['n'], TC_EGA),
    (T_701, C_EGA, 'vigente', EGA_DATA['701']['total_usd'], EGA_DATA['701']['anticipo'], SPLIT_A_EGA, EGA_DATA['701']['n'], TC_EGA),
    (T_202, C_202, 'vigente', TOTAL_USD_202, ANTICIPO_USD_202, SPLIT_A_202, N_CUOTAS_202, TC_202),
    (T_401, C_401, 'vigente', TOTAL_USD_401, ANTICIPO_USD_401, SPLIT_A_401, N_CUOTAS_401, TC_401),
]

rows_sql = []
for i, (tid, cid, est, prc, ant, splt, nq, tc) in enumerate(contrato_rows):
    sep = ',' if i < len(contrato_rows) - 1 else ''
    rows_sql.append(
        "    (obra_id, " + q(tid) + ", " + q(cid) + ", " + q(est) + ", " +
        q(prc) + ", " + q(ant) + ", " + q(splt) + ", " + str(nq) + ", " + q(tc) + ")" + sep
    )

# Fix: contratos table has obra_id as first positional
contratos_lines2 = [
    "-- insertar_contratos.sql",
    "-- Contratos de la obra Guatemala 5934",
    "-- IMPORTANTE: correr insertar_clientes.sql antes",
    "",
    "DO $$ DECLARE obra_uuid UUID;",
    "BEGIN",
    "  SELECT id INTO obra_uuid FROM obras WHERE nombre LIKE '%Guatemala%' LIMIT 1;",
    "",
    "  INSERT INTO contratos (id, obra_id, cliente_id, estado, precio_usd_total, anticipo_usd, split_a_pct, n_cuotas, tc_contrato) VALUES",
]
for i, (tid, cid, est, prc, ant, splt, nq, tc) in enumerate(contrato_rows):
    sep = ',' if i < len(contrato_rows) - 1 else ''
    contratos_lines2.append(
        "    (" + q(tid) + ", obra_uuid, " + q(cid) + ", " + q(est) + ", " +
        q(prc) + ", " + q(ant) + ", " + q(splt) + ", " + str(nq) + ", " + q(tc) + ")" + sep
    )
contratos_lines2.extend([
    "  ON CONFLICT (id) DO NOTHING;",
    "END $$;",
])

with open(os.path.join(OUTPUT, 'insertar_contratos.sql'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(contratos_lines2))
print("insertar_contratos.sql generado")

# ── 3. CUOTAS SQL ─────────────────────────────────────────────────────────────
wb_c = openpyxl.load_workbook(CUOTAS, data_only=True)

def find_next_section(rows, from_idx):
    """Devuelve el índice de la siguiente sección de cuota, o len(rows)."""
    for k in range(from_idx + 1, len(rows)):
        rs = str(rows[k])
        if 'CUOTA N' in rs and 'FECHA VENC' in rs:
            return k
    return len(rows)

def parse_cuotas_from_sheet(ws, col_n, col_fecha, cuota_a, cuota_b):
    rows = list(ws.iter_rows(values_only=True))
    result = []
    for i, row in enumerate(rows):
        rs = str(row)
        if 'CUOTA N' in rs and 'FECHA VENC' in rs:
            dr = rows[i+1] if i+1 < len(rows) else None
            if dr is None:
                continue
            n_v = dr[col_n] if col_n < len(dr) else None
            fecha_v = dr[col_fecha] if col_fecha < len(dr) else None
            if not is_num(n_v):
                continue
            # Buscar "cobrado"/"pagada" en el separador ANTES del header (hasta 6 filas)
            cobrado = False
            for j in range(max(0, i-6), i):
                if cobrado_in_row(rows[j]):
                    cobrado = True
                    break
            # Buscar "cobrado" dentro del cuerpo de la sección
            # Nos detenemos 4 filas antes del próximo header para evitar capturar separadores
            next_sec = find_next_section(rows, i)
            body_end = max(i+1, next_sec - 4)
            for k in range(i+1, min(body_end, len(rows))):
                if cobrado_in_row(rows[k]):
                    cobrado = True
            estado = 'pagada' if cobrado else 'pendiente'
            n_int = int(n_v)
            # Deduplicar: conservar la primera ocurrencia (la más reciente)
            if not any(c['n'] == n_int for c in result):
                result.append({'n': n_int, 'fecha': fecha_v, 'estado': estado, 'ca': cuota_a, 'cb': cuota_b})
    return result

def parse_ega_cuotas(ws):
    rows = list(ws.iter_rows(values_only=True))
    result = {k: [] for k in ('201', '301', '501', '701')}
    tid_map = {'201': T_201, '301': T_301, '501': T_501, '701': T_701}
    positions = [('201', 8, 9, 10), ('301', 21, 22, 23), ('501', 34, 35, 36), ('701', 47, 48, 49)]

    for i, row in enumerate(rows):
        rs = str(row)
        if 'CUOTA N' in rs and 'FECHA VENC' in rs and 'DEPTO' in rs:
            dr = rows[i+1] if i+1 < len(rows) else None
            if dr is None:
                continue
            cobrado = False
            for j in range(max(0, i-6), i):
                if cobrado_in_row(rows[j]):
                    cobrado = True
                    break
            next_sec = find_next_section(rows, i)
            body_end = max(i+1, next_sec - 4)
            for k in range(i+1, min(body_end, len(rows))):
                if cobrado_in_row(rows[k]):
                    cobrado = True
            estado = 'pagada' if cobrado else 'pendiente'
            for unit, dc, nc, fc in positions:
                if nc < len(dr) and is_num(dr[nc]):
                    fecha_v = dr[fc] if fc < len(dr) else None
                    ed = EGA_DATA[unit]
                    n_int = int(dr[nc])
                    if not any(c['n'] == n_int for c in result[unit]):
                        result[unit].append({'n': n_int, 'fecha': fecha_v, 'estado': estado,
                                            'ca': ed['cuota_a'], 'cb': ed['cuota_b']})
    return result

cuotas_801 = parse_cuotas_from_sheet(wb_c["CTA 801 SHARON Y MATI"], 5, 6, CUOTA_A_801, CUOTA_B_801)
cuotas_402 = parse_cuotas_from_sheet(wb_c["CTA 402 RUTH Y BENY"], 5, 6, CUOTA_A_402, CUOTA_B_402)
cuotas_202 = parse_cuotas_from_sheet(wb_c["CTA 202 GELLES"], 3, 4, CUOTA_A_202, CUOTA_B_202)
cuotas_401 = parse_cuotas_from_sheet(wb_c["CTA 401 ALBERTO I. Y GABRIELA H"], 5, 6, CUOTA_A_401, CUOTA_B_401)
cuotas_ega = parse_ega_cuotas(wb_c["CTA 201-301-501-701 EDGARDO Y R"])

print(f"801: {len(cuotas_801)} cuotas")
print(f"402: {len(cuotas_402)} cuotas")
print(f"202: {len(cuotas_202)} cuotas")
print(f"401: {len(cuotas_401)} cuotas")
for u, lst in cuotas_ega.items():
    print(f"{u}: {len(lst)} cuotas")

# Print estados
for label, lst in [('801', cuotas_801), ('402', cuotas_402), ('202', cuotas_202), ('401', cuotas_401)]:
    pagadas = sum(1 for c in lst if c['estado'] == 'pagada')
    pend = sum(1 for c in lst if c['estado'] == 'pendiente')
    print(f"  {label}: {pagadas} pagadas, {pend} pendientes")

all_cuota_sets = [
    (T_801, cuotas_801),
    (T_402, cuotas_402),
    (T_202, cuotas_202),
    (T_401, cuotas_401),
    (T_201, cuotas_ega['201']),
    (T_301, cuotas_ega['301']),
    (T_501, cuotas_ega['501']),
    (T_701, cuotas_ega['701']),
]

cuotas_lines = [
    "-- insertar_cuotas.sql",
    "-- Cuotas de cada contrato - obra Guatemala 5934",
    "-- IMPORTANTE: correr insertar_contratos.sql antes",
    "",
    "INSERT INTO cuotas (id, contrato_id, n_cuota, fecha_vencimiento, cuota_a_ars, cuota_b_ars,",
    "                    monto_pagado_a, monto_pagado_b, estado) VALUES",
]

cuota_rows_sql = []
for contrato_id, cuotas_list in all_cuota_sets:
    for c in cuotas_list:
        cid_q = q(gen_uuid())
        tid_q = q(contrato_id)
        n_q = str(c['n'])
        fecha_q = fmt_date(c['fecha'])
        ca_q = q(c['ca'])
        cb_q = q(c['cb'])
        pagado_a = q(0)
        pagado_b = q(0)
        estado_q = q(c['estado'])
        cuota_rows_sql.append(
            "  (" + cid_q + ", " + tid_q + ", " + n_q + ", " + fecha_q + ", " +
            ca_q + ", " + cb_q + ", " + pagado_a + ", " + pagado_b + ", " + estado_q + ")"
        )

for i, r in enumerate(cuota_rows_sql):
    sep = ',' if i < len(cuota_rows_sql) - 1 else ''
    cuotas_lines.append(r + sep)

cuotas_lines.append("ON CONFLICT (id) DO NOTHING;")

with open(os.path.join(OUTPUT, 'insertar_cuotas.sql'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(cuotas_lines))
print("insertar_cuotas.sql generado (" + str(len(cuota_rows_sql)) + " cuotas)")

# ── 4. MOVIMIENTOS CAJA SQL ───────────────────────────────────────────────────
wb_a = openpyxl.load_workbook(ADMIN, data_only=True)

def parse_caja_cash(ws):
    rows = list(ws.iter_rows(values_only=True))
    movs = []
    last_fecha = None
    for i, row in enumerate(rows):
        if i < 2:
            continue  # skip headers
        fecha_v = row[0]
        concepto_v = row[1]
        empresa_v = row[2]
        ing_usd = row[4]   # ingreso USD
        egr_usd = row[5]   # egreso USD (negative)
        tc_v = row[8]      # TC
        ing_ars = row[10]  # ingreso ARS
        egr_ars = row[11]  # egreso ARS (negative)

        # Update last known date
        if fecha_v is not None:
            last_fecha = fecha_v
        fecha_use = last_fecha

        # Skip rows with no amounts
        has_amount = any(is_num(x) and x != 0 for x in [ing_usd, egr_usd, ing_ars, egr_ars])
        if not has_amount:
            continue

        # Skip pure saldo rows (no concepto, no empresa, no usd amount)
        if concepto_v is None and empresa_v is None and not any(is_num(x) for x in [ing_usd, egr_usd]):
            continue

        fecha_sql = fmt_date(fecha_use)
        if fecha_sql == 'NULL':
            continue

        # Determine tipo from USD first, then ARS
        monto_usd = None
        monto_ars = None
        tipo = 'ingreso'

        if is_num(ing_usd) and ing_usd > 0:
            monto_usd = ing_usd
            tipo = 'ingreso'
        elif is_num(egr_usd) and egr_usd < 0:
            monto_usd = abs(egr_usd)
            tipo = 'egreso'

        if is_num(ing_ars) and ing_ars > 0:
            monto_ars = ing_ars
            if monto_usd is None:
                tipo = 'ingreso'
        elif is_num(egr_ars) and egr_ars < 0:
            monto_ars = abs(egr_ars)
            if monto_usd is None:
                tipo = 'egreso'

        concepto_clean = str(concepto_v).strip() if concepto_v else ''
        empresa_clean = str(empresa_v).strip() if empresa_v else None

        # Skip subtotal rows
        if not concepto_clean and empresa_clean is None and monto_usd is None and monto_ars is None:
            continue

        movs.append({
            'caja_id': CAJA_CASH,
            'fecha': fecha_sql,
            'tipo': tipo,
            'concepto': concepto_clean or '(sin concepto)',
            'contraparte': empresa_clean,
            'monto_usd': monto_usd,
            'monto_ars': monto_ars,
            'tc': tc_v if is_num(tc_v) else None,
        })
    return movs

def parse_caja_ars(ws, caja_id, col_ing_usd, col_egr_usd, col_ing_ars, col_egr_ars, col_tc=None):
    rows = list(ws.iter_rows(values_only=True))
    movs = []
    last_fecha = None
    for i, row in enumerate(rows):
        if i < 2:
            continue
        if len(row) <= max(col_ing_usd, col_egr_usd, col_ing_ars, col_egr_ars):
            continue

        fecha_v = row[0]
        concepto_v = row[1]
        empresa_v = row[2]
        ing_usd = row[col_ing_usd]
        egr_usd = row[col_egr_usd]
        ing_ars = row[col_ing_ars]
        egr_ars = row[col_egr_ars]
        tc_v = row[col_tc] if col_tc and col_tc < len(row) else None

        if fecha_v is not None:
            last_fecha = fecha_v
        fecha_use = last_fecha

        has_amount = any(is_num(x) and x != 0 for x in [ing_usd, egr_usd, ing_ars, egr_ars])
        if not has_amount:
            continue

        fecha_sql = fmt_date(fecha_use)
        if fecha_sql == 'NULL':
            continue

        monto_usd = None
        monto_ars = None
        tipo = 'ingreso'

        if is_num(ing_usd) and ing_usd > 0:
            monto_usd = ing_usd
            tipo = 'ingreso'
        elif is_num(egr_usd) and egr_usd < 0:
            monto_usd = abs(egr_usd)
            tipo = 'egreso'

        if is_num(ing_ars) and ing_ars > 0:
            monto_ars = ing_ars
            if monto_usd is None:
                tipo = 'ingreso'
        elif is_num(egr_ars) and egr_ars < 0:
            monto_ars = abs(egr_ars)
            if monto_usd is None:
                tipo = 'egreso'

        concepto_clean = str(concepto_v).strip() if concepto_v else ''
        empresa_clean = str(empresa_v).strip() if empresa_v else None

        if concepto_clean == '' and empresa_clean is None:
            continue

        movs.append({
            'caja_id': caja_id,
            'fecha': fecha_sql,
            'tipo': tipo,
            'concepto': concepto_clean or '(sin concepto)',
            'contraparte': empresa_clean,
            'monto_usd': monto_usd,
            'monto_ars': monto_ars,
            'tc': tc_v if is_num(tc_v) else None,
        })
    return movs

movs_cash  = parse_caja_cash(wb_a["CAJA CASH"])
movs_banco = parse_caja_ars(wb_a["CAJA BANCO"], CAJA_BANCO, col_ing_usd=3, col_egr_usd=4, col_ing_ars=7, col_egr_ars=8)
movs_mn    = parse_caja_ars(wb_a["CAJA MN"],    CAJA_MN,    col_ing_usd=3, col_egr_usd=4, col_ing_ars=7, col_egr_ars=8)

print(f"CAJA CASH: {len(movs_cash)} movimientos")
print(f"CAJA BANCO: {len(movs_banco)} movimientos")
print(f"CAJA MN: {len(movs_mn)} movimientos")

all_movs = movs_cash + movs_banco + movs_mn

mov_lines = [
    "-- insertar_movimientos_caja.sql",
    "-- Movimientos de CAJA CASH, CAJA BANCO y CAJA MN - obra Guatemala 5934",
    "-- IMPORTANTE: las cajas deben existir previamente (ver insertar_cajas.sql al final)",
    "",
    "-- Paso 1: Crear cajas si no existen",
    "DO $$ DECLARE obra_uuid UUID;",
    "BEGIN",
    "  SELECT id INTO obra_uuid FROM obras WHERE nombre LIKE '%Guatemala%' LIMIT 1;",
    "",
    "  INSERT INTO cajas (id, obra_id, nombre) VALUES",
    "    (" + q(CAJA_CASH)  + ", obra_uuid, 'CAJA CASH'),",
    "    (" + q(CAJA_BANCO) + ", obra_uuid, 'CAJA BANCO'),",
    "    (" + q(CAJA_MN)    + ", obra_uuid, 'CAJA MN')",
    "  ON CONFLICT (id) DO NOTHING;",
    "END $$;",
    "",
    "-- Paso 2: Insertar movimientos",
    "INSERT INTO movimientos_caja",
    "  (id, caja_id, obra_id, fecha, tipo, concepto, contraparte, monto_usd, monto_ars, tc_blue, origen)",
    "SELECT",
    "  gen_random_uuid(), m.caja_id, o.id, m.fecha::date, m.tipo,",
    "  m.concepto, m.contraparte, m.monto_usd, m.monto_ars, m.tc_blue, 'importacion'",
    "FROM (VALUES",
]

mov_val_rows = []
obra_subq = "(SELECT id FROM obras WHERE nombre LIKE '%Guatemala%' LIMIT 1)"
for m in all_movs:
    row = (
        "  (" +
        q(m['caja_id']) + ", " +
        m['fecha'] + ", " +
        q(m['tipo']) + ", " +
        q(m['concepto']) + ", " +
        q(m['contraparte']) + ", " +
        (q(round(m['monto_usd'], 2)) if m['monto_usd'] is not None else 'NULL') + "::numeric, " +
        (q(round(m['monto_ars'], 2)) if m['monto_ars'] is not None else 'NULL') + "::numeric, " +
        (q(m['tc']) if m['tc'] is not None else 'NULL') + "::numeric" +
        ")"
    )
    mov_val_rows.append(row)

for i, r in enumerate(mov_val_rows):
    sep = ',' if i < len(mov_val_rows) - 1 else ''
    mov_lines.append(r + sep)

mov_lines.extend([
    ") AS m(caja_id, fecha, tipo, concepto, contraparte, monto_usd, monto_ars, tc_blue)",
    "CROSS JOIN (SELECT id FROM obras WHERE nombre LIKE '%Guatemala%' LIMIT 1) o",
    "ON CONFLICT DO NOTHING;",
])

with open(os.path.join(OUTPUT, 'insertar_movimientos_caja.sql'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(mov_lines))
print("insertar_movimientos_caja.sql generado (" + str(len(all_movs)) + " movimientos)")

print("")
print("=== RESUMEN ===")
print("Archivos generados en: " + OUTPUT)
print("  insertar_clientes.sql    - 5 clientes")
print("  insertar_contratos.sql   - 8 contratos (801, 402, 201, 301, 501, 701, 202, 401)")
print("  insertar_cuotas.sql      - " + str(len(cuota_rows_sql)) + " cuotas")
print("  insertar_movimientos_caja.sql - " + str(len(all_movs)) + " movimientos")
